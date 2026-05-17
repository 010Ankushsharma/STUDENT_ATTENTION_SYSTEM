
"""
dashboard/api.py
FastAPI application with all REST endpoints and WebSocket handler.

API Endpoints:
    GET  /                        â†’ Dashboard HTML page
    GET  /api/status              â†’ System status
    GET  /api/session/current     â†’ Current session info
    GET  /api/students            â†’ All student scores
    GET  /api/students/{id}       â†’ Single student detail
    GET  /api/alerts              â†’ Recent alerts
    GET  /api/analytics/overview  â†’ Session analytics
    GET  /api/analytics/timeline  â†’ Score timeline data
    GET  /api/analytics/ranking   â†’ Student ranking
    GET  /api/analytics/distribution â†’ State distribution
    GET  /api/export/csv          â†’ Download CSV report
    GET  /api/export/excel        â†’ Download Excel report
    GET  /video_feed              â†’ MJPEG live stream
    WS   /ws                      â†’ Real-time WebSocket updates
"""
import os
import csv
import io
import json
import time
import asyncio
import threading
from datetime import datetime
from typing import Optional, Dict, List
from pathlib import Path

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Response
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from .config import dashboard_cfg

# Will be set by DashboardServer
_scoring_pipeline = None
_db_manager = None
_analytics_engine = None
_live_logger = None
_frame_source = None  # Callable that returns latest annotated frame (JPEG bytes)


def create_app(scoring=None, db=None, analytics=None,
               logger=None, frame_source=None) -> FastAPI:
    """
    Create the FastAPI application.

    Args:
        scoring:      ScoringPipeline instance
        db:           DatabaseManager instance
        analytics:    AnalyticsEngine instance
        logger:       LiveLogger instance
        frame_source: Callable returning latest JPEG frame bytes
    """
    global _scoring_pipeline, _db_manager, _analytics_engine
    global _live_logger, _frame_source

    _scoring_pipeline = scoring
    _db_manager = db
    _analytics_engine = analytics
    _live_logger = logger
    _frame_source = frame_source

    app = FastAPI(
        title="Student Attention Dashboard",
        description="Real-time classroom attention monitoring",
        version="1.0.0",
    )

    # CORS
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_methods=["*"],
        allow_headers=["*"],
    )

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ HTML Dashboard â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/", response_class=HTMLResponse)
    async def dashboard_page():
        """Serve the main dashboard HTML page."""
        html_path = Path(__file__).parent / "templates" / "index.html"
        if html_path.exists():
            return HTMLResponse(content=html_path.read_text(encoding="utf-8"), status_code=200)
        return HTMLResponse(content="<h1>Dashboard template not found</h1>", status_code=500)

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ System Status â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/api/status")
    async def get_status():
        """System status and health check."""
        session_id = _live_logger.session_id if _live_logger else None
        return {
            "status": "running",
            "session_active": session_id is not None,
            "session_id": session_id,
            "timestamp": datetime.now().isoformat(),
            "db_size_mb": _db_manager.get_db_size_mb() if _db_manager else 0,
            "table_counts": _db_manager.get_table_counts() if _db_manager else {},
        }

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Session â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/api/session/current")
    async def get_current_session():
        """Get current active session details."""
        if not _scoring_pipeline:
            return {"error": "No scoring pipeline"}
        summary = _scoring_pipeline.get_class_summary()
        session_id = _live_logger.session_id if _live_logger else None
        return {
            "session_id": session_id,
            "class_summary": summary,
            "leaderboard": _scoring_pipeline.get_leaderboard()[:dashboard_cfg.leaderboard_size],
            "total_alerts": len(_scoring_pipeline.get_alerts()),
        }

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Students â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/api/students")
    async def get_all_students():
        """Get current attention data for all students."""
        if not _scoring_pipeline:
            return {"students": []}
        students = _scoring_pipeline.get_all_students()
        result = []
        for sid, record in students.items():
            result.append({
                "student_id": sid,
                "score": record.current_score,
                "score_pct": round(record.current_score * 100, 1),
                "state": record.current_state.value,
                "attention_pct": record.attention_percentage,
                "distribution": record.state_distribution,
                "recent_avg": record.recent_score_avg,
                "total_frames": record.total_frames,
            })
        return {"students": result, "count": len(result)}

    @app.get("/api/students/{student_id}")
    async def get_student(student_id: int):
        """Get detailed data for a specific student."""
        if not _scoring_pipeline:
            return {"error": "No pipeline"}
        record = _scoring_pipeline.get_student(student_id)
        if not record:
            return {"error": f"Student {student_id} not found"}

        # Get DB history if available
        history = []
        if _analytics_engine and _live_logger and _live_logger.session_id:
            summary = _analytics_engine.student_summary(
                _live_logger.session_id, student_id
            )
        else:
            summary = {}

        return {
            "student_id": student_id,
            "score": record.current_score,
            "score_pct": round(record.current_score * 100, 1),
            "state": record.current_state.value,
            "attention_pct": record.attention_percentage,
            "distribution": record.state_distribution,
            "total_frames": record.total_frames,
            "db_summary": summary,
        }

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Alerts â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/api/alerts")
    async def get_alerts():
        """Get recent alerts."""
        if not _scoring_pipeline:
            return {"alerts": []}
        alerts = _scoring_pipeline.get_alerts()
        result = []
        for a in alerts[-dashboard_cfg.alert_display_count:]:
            result.append({
                "student_id": a.student_id,
                "type": a.alert_type.value,
                "severity": a.severity.value,
                "message": a.message,
                "score": a.score,
                "timestamp": a.timestamp,
                "time_ago": f"{time.time() - a.timestamp:.0f}s ago",
            })
        return {"alerts": list(reversed(result)), "total": len(alerts)}

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Analytics â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/api/analytics/overview")
    async def get_analytics_overview():
        """Session analytics overview."""
        if not _analytics_engine or not _live_logger:
            return {"error": "Analytics not available"}
        sid = _live_logger.session_id
        if not sid:
            return {"error": "No active session"}
        _live_logger.flush()
        return _analytics_engine.session_overview(sid)

    @app.get("/api/analytics/timeline")
    async def get_analytics_timeline(bucket_seconds: int = 10):
        """Score timeline for charts."""
        if not _analytics_engine or not _live_logger:
            return {"timeline": []}
        sid = _live_logger.session_id
        if not sid:
            return {"timeline": []}
        _live_logger.flush()
        data = _analytics_engine.attention_timeline(sid, bucket_seconds)
        return {"timeline": data, "bucket_seconds": bucket_seconds}

    @app.get("/api/analytics/ranking")
    async def get_analytics_ranking():
        """Student ranking by attention."""
        if not _analytics_engine or not _live_logger:
            return {"ranking": []}
        sid = _live_logger.session_id
        if not sid:
            return {"ranking": []}
        _live_logger.flush()
        return {"ranking": _analytics_engine.class_ranking(sid)}

    @app.get("/api/analytics/distribution")
    async def get_analytics_distribution():
        """State distribution percentages."""
        if not _analytics_engine or not _live_logger:
            return {"distribution": {}}
        sid = _live_logger.session_id
        if not sid:
            return {"distribution": {}}
        _live_logger.flush()
        return {"distribution": _analytics_engine.state_distribution(sid)}

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Export â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/api/export/csv")
    async def export_csv():
        """Export session data as CSV."""
        if not _analytics_engine or not _live_logger:
            return Response("No data", status_code=404)
        sid = _live_logger.session_id
        if not sid:
            return Response("No active session", status_code=404)
        _live_logger.flush()

        ranking = _analytics_engine.class_ranking(sid)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Rank", "Student ID", "Avg Score", "Attentive %",
                         "Total Frames"])
        for r in ranking:
            writer.writerow([
                r["rank"], r["student_id"],
                r["avg_score"], r["attentive_pct"], r["total_frames"]
            ])

        content = output.getvalue()
        filename = f"attention_report_{sid}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.get("/api/export/excel")
    async def export_excel():
        """Export session data as Excel (via openpyxl if available, else CSV)."""
        if not _analytics_engine or not _live_logger:
            return Response("No data", status_code=404)
        sid = _live_logger.session_id
        if not sid:
            return Response("No active session", status_code=404)
        _live_logger.flush()

        try:
            from openpyxl import Workbook
            wb = Workbook()

            # Sheet 1: Ranking
            ws1 = wb.active
            ws1.title = "Student Ranking"
            ws1.append(["Rank", "Student ID", "Avg Score", "Attentive %", "Total Frames"])
            for r in _analytics_engine.class_ranking(sid):
                ws1.append([r["rank"], r["student_id"], r["avg_score"],
                           r["attentive_pct"], r["total_frames"]])

            # Sheet 2: State Distribution
            ws2 = wb.create_sheet("State Distribution")
            ws2.append(["State", "Percentage"])
            for state, pct in _analytics_engine.state_distribution(sid).items():
                ws2.append([state, pct])

            # Sheet 3: Timeline
            ws3 = wb.create_sheet("Timeline")
            ws3.append(["Student ID", "Time Bucket", "Avg Score", "Min", "Max", "Samples"])
            for t in _analytics_engine.attention_timeline(sid, 30):
                ws3.append([t["student_id"], t["time_bucket"],
                           t["avg_score"], t["min_score"], t["max_score"],
                           t["samples"]])

            # Save to bytes
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            filename = f"attention_report_{sid}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            return Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            # Fallback to CSV
            return await export_csv()

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Video Feed (MJPEG) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @app.get("/video_feed")
    async def video_feed():
        """MJPEG stream of the annotated webcam feed."""
        async def generate():
            while True:
                if _frame_source:
                    frame_bytes = _frame_source()
                    if frame_bytes:
                        yield (b"--frame\r\n"
                               b"Content-Type: image/jpeg\r\n\r\n"
                               + frame_bytes + b"\r\n")
                await asyncio.sleep(1.0 / dashboard_cfg.stream_fps)

        return StreamingResponse(
            generate(),
            media_type="multipart/x-mixed-replace; boundary=frame"
        )

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ WebSocket (Real-time Updates) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    connected_clients: List[WebSocket] = []

    @app.websocket("/ws")
    async def websocket_endpoint(ws: WebSocket):
        """WebSocket for real-time score updates."""
        await ws.accept()
        connected_clients.append(ws)
        try:
            while True:
                # Push update to client
                if _scoring_pipeline:
                    summary = _scoring_pipeline.get_class_summary()
                    students = []
                    for sid, record in _scoring_pipeline.get_all_students().items():
                        students.append({
                            "id": sid,
                            "score": round(record.current_score, 3),
                            "state": record.current_state.value,
                            "pct": record.attention_percentage,
                        })

                    alerts = _scoring_pipeline.get_alerts()
                    recent_alerts = []
                    for a in alerts[-5:]:
                        recent_alerts.append({
                            "student_id": a.student_id,
                            "type": a.alert_type.value,
                            "severity": a.severity.value,
                            "message": a.message,
                        })

                    payload = {
                        "type": "update",
                        "timestamp": time.time(),
                        "class_summary": summary,
                        "students": students,
                        "recent_alerts": recent_alerts,
                    }
                    await ws.send_json(payload)

                await asyncio.sleep(dashboard_cfg.ws_update_interval_ms / 1000.0)
        except WebSocketDisconnect:
            connected_clients.remove(ws)
        except Exception:
            if ws in connected_clients:
                connected_clients.remove(ws)

    return app


